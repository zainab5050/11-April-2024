# first install openpyxl then import it
import openpyxl

path = "D:\\zzz\\Book1.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook["Sheet2"]

rows = sheet.max_row
col = sheet.max_column

print(rows)
print(col)

for r in range(1, rows + 1):
    for c in range(1, col + 1):
        print(sheet.cell(row=r, column=c).value, end="   ")
    print()


# write data in Excel
for r in range(1, 5):
    for c in range(1, 4):
        sheet.cell(row=r, column=c).value = "wellcome"
    workbook.save(path)

# Data Driven Test
